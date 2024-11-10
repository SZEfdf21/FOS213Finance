import pandas as pd
import openpyxl as opx
import numpy as np


def schuldenUpdater(jaarrekening, wbName, wsName, code):
    # Correcte pagina schulden excel ophalen
    wb = opx.load_workbook(wbName, data_only=False)
    ws = wb[wsName]
    aantal = len(ws['A']) - 1

    # Betalingen bovenhalen
    jaarrekeningDf = pd.read_excel(jaarrekening, index_col=None, header=0, sheet_name="Jaarrekening")
    jaarrekeningDf = jaarrekeningDf.fillna(0)

    # Lijst maken met alle leiding en stam + Rekening In op 0 zetten en te vedelen waarden op 0 zetten
    # (want systeem bekijkt de hele jaarrekening):
    leidingLijst = []
    counter = 2
    while counter < (aantal + 2):
        leidingLijst.append(str(ws['B' + str(counter)].value).lower())
        ws['L' + str(counter)] = 0.0
        counter += 1

    # Checken in de jaarrekening welke betalingen van welke leiding zijn
    for index, row in jaarrekeningDf.iterrows():
        if row['Code'] == code:
            mededeling = str(row['Mededeling']).split()
            naam = mededeling[-1].lower()
            if naam in leidingLijst:
                rowPos = str(leidingLijst.index(naam) + 2)
                # Uitgaven zitten hierbij voor moest er is iets teveel betaald zijn
                ws['L' + rowPos] = ws['L' + rowPos].value + row['In'] + row['Uit']
                ws['L' + rowPos].number_format = '#,##0.00 €'

    wb.save(filename=wbName)
    return


def lidgeldInschrijvingUpdater(jaarrekening, wsName, code, naamType, betaaldCol):
    wb = opx.load_workbook(jaarrekening)
    ws = wb[wsName]

    jaarrekeningDf = pd.read_excel(jaarrekening, index_col=None, header=0, sheet_name="Jaarrekening")
    lidgeldInschrijvingDf = pd.read_excel(jaarrekening, header=0, sheet_name="LidgeldInschrijvingen")

    # Weghalen van de rijen zonder voor of achternaam + nan waarden naar lege string + index kolom maken
    lidgeldInschrijvingDf = lidgeldInschrijvingDf[lidgeldInschrijvingDf['Voornaam'].notna()]
    lidgeldInschrijvingDf = lidgeldInschrijvingDf[lidgeldInschrijvingDf['Achternaam'].notna()]
    lidgeldInschrijvingDf.fillna(0, inplace=True)
    lidgeldInschrijvingDf.reset_index(inplace=True)

    # Lijst maken met al de mededelingen en bijhorende bedragen die code bevatten, enkel naar IN kijken
    mededelingen = jaarrekeningDf[jaarrekeningDf['Code'] == code]['Mededeling'].tolist()
    bedragen = jaarrekeningDf[jaarrekeningDf['Code'] == code]['In'].tolist()
    # Dataframe naar 2d lijst:
    lidgeldInschrijvingLijst = lidgeldInschrijvingDf.to_numpy().tolist()

    # Kijken welke type het is
    lijstPos = 25
    if naamType == 'Lidgeld':
        lijstPos = 6
    elif naamType == 'EHWE':
        lijstPos = 8
    elif naamType == 'Kamp':
        lijstPos = 10

    # Alle leden overlopen
    for row in lidgeldInschrijvingLijst:
        # Kijken of dat iemand al betaald heeft
        if row[lijstPos] != 1:
            voornaam = row[2].lower()
            achternaam = row[3].lower()
            # Kijken of er een totem is meegegeven in de ledenlijst:
            totem = str(row[4]).lower()
            if row[4] == 0:
                totem = 'geenTotemMegegeven'

            # Kijken of het bedrag 0 euro is:
            if row[(lijstPos-1)] == 0:
                row[lijstPos] = 1

            # Kijken in de mededelingen of lid heeft betaald:
            i = 0
            found = False
            while i < len(mededelingen) and found == False:
                # Kijken of de voor en achternaam er in staat (en kijken of er een totem in staat):
                mededeling = str(mededelingen[i]).lower()
                if (voornaam in mededeling and achternaam in mededeling) or (totem in mededeling):
                    found = True
                    # Kijken of het bedrag klopt:
                    if float(bedragen[i]) == float(row[(lijstPos-1)]):
                        # Waarde 1 voor correct bedrag invullen:
                        row[lijstPos] = 1
                    # Waarde 2 voor foutief bedrag invullen:
                    else:
                        row[lijstPos] = 2
                i += 1

            # Aanpassen in de Excel
            excelIndex = str(row[0] + 2)
            ws[betaaldCol + excelIndex] = row[lijstPos]
    wb.save(filename=jaarrekening)
    return


if __name__ == '__main__':
    schuldenoverzicht = "Excels/2324SchuldenOverzicht.xlsx"
    jaarrekening = "Excels/2324Jaarrekening.xlsx"

    # schuldenUpdater(jaarrekening, schuldenoverzicht, "Inleefweek", 430)
    # schuldenUpdater(jaarrekening, schuldenoverzicht, "LW", 420)
    # schuldenUpdater(jaarrekening, schuldenoverzicht, "Voorschotten allerlei", 410)
    # schuldenUpdater(jaarrekening, schuldenoverzicht, "Vorig jaar", 910)
    schuldenUpdater(jaarrekening, schuldenoverzicht, "TeBetalen", 300)

    # Voor lidgelden te controleren (locatie jaarrekening, sheetNaam, code, naam van het type, kolom in excel):
    # lidgeldInschrijvingUpdater(jaarrekening, "LidgeldInschrijvingen", 100, 'Lidgeld', 'F')

    # Voor EHWE te controleren (locatie jaarrekening, sheetNaam, code, naam van het type, kolom in excel):
    # lidgeldInschrijvingUpdater(jaarrekening, "LidgeldInschrijvingen", 505, 'EHWE', 'H')

    # Voor kamp te controleren (locatie jaarrekening, sheetNaam, code, naam van het type, kolom in excel):
    # lidgeldInschrijvingUpdater(jaarrekening, "LidgeldInschrijvingen", 605, 'Kamp', 'J')

